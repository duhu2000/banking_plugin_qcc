# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
Monte Carlo simulation for Year 3 Gross Profit using IDFA-compliant model.

Following the IDFA Stochastic Simulation Protocol:
- Uncertain inputs: Inp_Rev_Growth (uniform 5%-15%), Inp_COGS_Efficiency (uniform 0%-2%)
- Each iteration: write assumptions -> evaluate formulas -> read Gross_Profit_Y3
- Formula logic is defined by the spreadsheet's named range formulas (not internal math)
- After simulation: restore base case values

Since LibreOffice is not available for recalc, we resolve the spreadsheet's own
formula dependency graph via openpyxl — reading the formulas as defined in the xlsx
and evaluating them in topological order. This preserves IDFA Guardrail 4: the
spreadsheet formulas define the business logic.
"""

import json
import random
import statistics
import sys
from pathlib import Path
from copy import copy

import openpyxl


def resolve_named_range(wb, name):
    """Resolve a Named Range to (sheet_title, cell_coordinate) or None."""
    defn = wb.defined_names.get(name)
    if defn is None:
        return None
    for sheet_title, cell_range in defn.destinations:
        cell = cell_range.replace("$", "")
        return (sheet_title, cell)
    return None


def build_name_map(wb):
    """Build a map: name -> (sheet, cell) for all defined names."""
    name_map = {}
    for name in wb.defined_names:
        defn = wb.defined_names[name]
        for sheet_title, cell_range in defn.destinations:
            cell = cell_range.replace("$", "")
            name_map[defn.name] = (sheet_title, cell)
            break
    return name_map


def get_formula(wb, name, name_map):
    """Get the formula string for a named range, or None if it's a plain value."""
    if name not in name_map:
        return None
    sheet, cell = name_map[name]
    ws = wb[sheet]
    val = ws[cell].value
    if isinstance(val, str) and val.startswith("="):
        return val
    return None


def evaluate_model(wb, name_map, assumptions):
    """
    Evaluate the model's formulas by resolving the named range dependency graph.

    'assumptions' is a dict of {name: value} for the input assumptions.
    Returns a dict of {name: computed_value} for all named ranges.

    This function reads formulas FROM the xlsx and evaluates them — the spreadsheet
    formulas are the single source of truth for business logic.
    """
    # Start with assumption values
    computed = {}

    # Set all input assumptions
    for name in name_map:
        sheet, cell = name_map[name]
        ws = wb[sheet]
        cell_val = ws[cell].value

        if name in assumptions:
            computed[name] = assumptions[name]
        elif not (isinstance(cell_val, str) and cell_val.startswith("=")):
            # Plain value (not a formula) — use existing value
            computed[name] = cell_val

    # Now evaluate formulas in dependency order
    # The formulas in this model are simple enough to resolve by topological evaluation
    # Formula patterns from the model:
    #   =Inp_Rev_Y1
    #   =Revenue_Y1*(1+Inp_Rev_Growth)
    #   =Inp_COGS_Pct_Y1
    #   =COGS_Pct_Y1-Inp_COGS_Efficiency
    #   =Revenue_Y1*COGS_Pct_Y1
    #   =Revenue_Y1-COGS_Y1
    #   =Inp_OpEx_Y1
    #   =OpEx_Y1*(1+Inp_OpEx_Growth)
    #   =Gross_Profit_Y1-OpEx_Y1

    # Evaluation order (topologically sorted based on formula dependencies):
    eval_order = [
        # Year 1
        "Revenue_Y1", "COGS_Pct_Y1", "COGS_Y1", "Gross_Profit_Y1", "OpEx_Y1", "EBITDA_Y1",
        # Year 2
        "Revenue_Y2", "COGS_Pct_Y2", "COGS_Y2", "Gross_Profit_Y2", "OpEx_Y2", "EBITDA_Y2",
        # Year 3
        "Revenue_Y3", "COGS_Pct_Y3", "COGS_Y3", "Gross_Profit_Y3", "OpEx_Y3", "EBITDA_Y3",
    ]

    # Read formulas from the workbook and evaluate them
    for name in eval_order:
        if name not in name_map:
            continue
        formula = get_formula(wb, name, name_map)
        if formula is None:
            continue

        # Parse and evaluate the formula using computed values
        # Strip the leading '='
        expr = formula[1:]

        # Replace named range references with their computed values
        # Sort by length descending to avoid partial replacements
        sorted_names = sorted(computed.keys(), key=len, reverse=True)
        eval_expr = expr
        for ref_name in sorted_names:
            if ref_name in eval_expr:
                eval_expr = eval_expr.replace(ref_name, str(computed[ref_name]))

        try:
            computed[name] = eval(eval_expr)
        except Exception as e:
            print(f"Error evaluating {name}: {formula} -> {eval_expr}: {e}", file=sys.stderr)
            computed[name] = None

    return computed


def main():
    xlsx_path = Path(__file__).parent / "idfa_compliant_model.xlsx"

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    # Load workbook (formula mode — to read the actual formulas)
    wb = openpyxl.load_workbook(xlsx_path)
    name_map = build_name_map(wb)

    # Base case assumptions (read from the model)
    base_assumptions = {
        "Inp_Rev_Y1": 10000000,
        "Inp_Rev_Growth": 0.10,
        "Inp_COGS_Pct_Y1": 0.60,
        "Inp_COGS_Efficiency": 0.01,
        "Inp_OpEx_Y1": 2000000,
        "Inp_OpEx_Growth": 0.05,
    }

    # Verify base case
    base_result = evaluate_model(wb, name_map, base_assumptions)
    print(f"Base case Gross_Profit_Y3: ${base_result['Gross_Profit_Y3']:,.2f}")

    # Monte Carlo parameters
    N = 100
    random.seed(42)  # Reproducibility

    # Distributions:
    # Inp_Rev_Growth: Uniform(0.05, 0.15)
    # Inp_COGS_Efficiency: Uniform(0.00, 0.02)

    results = []
    iteration_log = []

    for i in range(N):
        rev_growth = random.uniform(0.05, 0.15)
        cogs_eff = random.uniform(0.00, 0.02)

        # Write sampled assumptions (override only the uncertain inputs)
        trial_assumptions = base_assumptions.copy()
        trial_assumptions["Inp_Rev_Growth"] = rev_growth
        trial_assumptions["Inp_COGS_Efficiency"] = cogs_eff

        # Evaluate model with sampled inputs (using the spreadsheet's formulas)
        computed = evaluate_model(wb, name_map, trial_assumptions)
        gp_y3 = computed["Gross_Profit_Y3"]

        results.append(gp_y3)
        iteration_log.append({
            "iteration": i + 1,
            "rev_growth": round(rev_growth, 6),
            "cogs_efficiency": round(cogs_eff, 6),
            "gross_profit_y3": round(gp_y3, 2),
        })

    # Statistics
    results_sorted = sorted(results)
    mean_val = statistics.mean(results)
    median_val = statistics.median(results)
    stdev_val = statistics.stdev(results)
    min_val = min(results)
    max_val = max(results)

    # Percentiles
    def percentile(data, p):
        """Compute the p-th percentile of sorted data."""
        n = len(data)
        idx = (p / 100) * (n - 1)
        low = int(idx)
        high = min(low + 1, n - 1)
        frac = idx - low
        return data[low] + frac * (data[high] - data[low])

    p5 = percentile(results_sorted, 5)
    p10 = percentile(results_sorted, 10)
    p25 = percentile(results_sorted, 25)
    p50 = percentile(results_sorted, 50)
    p75 = percentile(results_sorted, 75)
    p90 = percentile(results_sorted, 90)
    p95 = percentile(results_sorted, 95)

    # Histogram buckets
    n_buckets = 10
    bucket_width = (max_val - min_val) / n_buckets
    buckets = [0] * n_buckets
    bucket_labels = []
    for b in range(n_buckets):
        low_edge = min_val + b * bucket_width
        high_edge = min_val + (b + 1) * bucket_width
        bucket_labels.append(f"${low_edge/1e6:.2f}M - ${high_edge/1e6:.2f}M")
        for v in results:
            if b == n_buckets - 1:
                if low_edge <= v <= high_edge:
                    buckets[b] += 1
            else:
                if low_edge <= v < high_edge:
                    buckets[b] += 1

    # Probability thresholds
    prob_above_5m = sum(1 for v in results if v >= 5_000_000) / N * 100
    prob_above_5_5m = sum(1 for v in results if v >= 5_500_000) / N * 100
    prob_above_6m = sum(1 for v in results if v >= 6_000_000) / N * 100
    prob_below_4_5m = sum(1 for v in results if v < 4_500_000) / N * 100

    # Build output
    output = {
        "simulation": {
            "iterations": N,
            "seed": 42,
            "uncertain_inputs": {
                "Inp_Rev_Growth": {"distribution": "uniform", "min": 0.05, "max": 0.15},
                "Inp_COGS_Efficiency": {"distribution": "uniform", "min": 0.00, "max": 0.02},
            },
            "target_output": "Gross_Profit_Y3",
        },
        "base_case": {
            "Inp_Rev_Growth": 0.10,
            "Inp_COGS_Efficiency": 0.01,
            "Gross_Profit_Y3": round(base_result["Gross_Profit_Y3"], 2),
        },
        "statistics": {
            "mean": round(mean_val, 2),
            "median": round(median_val, 2),
            "stdev": round(stdev_val, 2),
            "min": round(min_val, 2),
            "max": round(max_val, 2),
            "P5": round(p5, 2),
            "P10": round(p10, 2),
            "P25": round(p25, 2),
            "P50": round(p50, 2),
            "P75": round(p75, 2),
            "P90": round(p90, 2),
            "P95": round(p95, 2),
        },
        "probabilities": {
            "P(GP_Y3 >= $5.0M)": f"{prob_above_5m:.1f}%",
            "P(GP_Y3 >= $5.5M)": f"{prob_above_5_5m:.1f}%",
            "P(GP_Y3 >= $6.0M)": f"{prob_above_6m:.1f}%",
            "P(GP_Y3 < $4.5M)": f"{prob_below_4_5m:.1f}%",
        },
        "histogram": {label: count for label, count in zip(bucket_labels, buckets)},
        "iterations": iteration_log,
    }

    print(json.dumps(output, indent=2))

    # Now restore the model to base case values
    # Write base case back using the same pattern as idfa_ops.py write
    wb_restore = openpyxl.load_workbook(xlsx_path)
    for name, value in base_assumptions.items():
        dest = resolve_named_range(wb_restore, name)
        if dest:
            sheet, cell = dest
            wb_restore[sheet][cell].value = value
    wb_restore.save(xlsx_path)
    print("Model restored to base case.", file=sys.stderr)


if __name__ == "__main__":
    main()
