"""
Generate three XLSX eval files for IDFA financial architecture skill evaluations.

File 1: idfa_compliant_model.xlsx  - IDFA-compliant 3-year gross profit waterfall
File 2: legacy_model.xlsx          - Legacy model with deliberate IDFA violations
File 3: ops_test_model.xlsx        - Model for testing idfa-ops script operations
"""

import openpyxl
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import quote_sheetname

BASE_DIR_FA = "financial-architect/evals/files"
BASE_DIR_OPS = "idfa-ops/evals/files"

# ─────────────────────────────────────────────────────────────────────────────
# Helper: define a named range scoped to a workbook
# ─────────────────────────────────────────────────────────────────────────────

def add_named_range(wb, name, sheet_title, cell_coord):
    """Add a workbook-scoped named range pointing to sheet!cell."""
    # Build fully absolute reference like $B$2
    col = "".join(c for c in cell_coord if c.isalpha())
    row = "".join(c for c in cell_coord if c.isdigit())
    abs_addr = f"${col}${row}"
    ref = f"{quote_sheetname(sheet_title)}!{abs_addr}"

    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)


def style_header(ws, row, cols, text_list):
    """Apply header styling to a row."""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    for i, col in enumerate(cols):
        cell = ws.cell(row=row, column=col, value=text_list[i])
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


# ═════════════════════════════════════════════════════════════════════════════
# FILE 1: IDFA-Compliant Model
# ═════════════════════════════════════════════════════════════════════════════

def create_compliant_model():
    wb = openpyxl.Workbook()

    # ── Sheet: Assumptions ───────────────────────────────────────────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.column_dimensions["A"].width = 30
    ws_a.column_dimensions["B"].width = 18

    style_header(ws_a, 1, [1, 2], ["Assumption", "Value"])

    assumptions = [
        ("Inp_Rev_Y1",          "Revenue Year 1",          10_000_000, "B2"),
        ("Inp_Rev_Growth",      "Revenue Growth Rate",      0.10,       "B3"),
        ("Inp_COGS_Pct_Y1",    "COGS % Year 1",           0.60,       "B4"),
        ("Inp_COGS_Efficiency", "COGS Efficiency Gain/yr", 0.01,       "B5"),
        ("Inp_OpEx_Y1",        "OpEx Year 1",              2_000_000,  "B6"),
        ("Inp_OpEx_Growth",    "OpEx Growth Rate",          0.05,       "B7"),
    ]

    for idx, (name, label, value, cell) in enumerate(assumptions, start=2):
        ws_a.cell(row=idx, column=1, value=label)
        ws_a.cell(row=idx, column=2, value=value)
        ws_a[cell].number_format = '#,##0.00' if isinstance(value, float) and value < 1 else '#,##0'
        add_named_range(wb, name, "Assumptions", cell)

    # ── Sheet: Calculations ──────────────────────────────────────────────
    ws_c = wb.create_sheet("Calculations")
    ws_c.column_dimensions["A"].width = 22
    ws_c.column_dimensions["B"].width = 18
    ws_c.column_dimensions["C"].width = 18
    ws_c.column_dimensions["D"].width = 18

    style_header(ws_c, 1, [1, 2, 3, 4], ["Metric", "Year 1", "Year 2", "Year 3"])

    # Layout:
    # Row 2: Revenue
    # Row 3: COGS %
    # Row 4: COGS $
    # Row 5: Gross Profit
    # Row 6: OpEx
    # Row 7: EBITDA

    calc_rows = {
        "Revenue":      2,
        "COGS_Pct":     3,
        "COGS":         4,
        "Gross_Profit": 5,
        "OpEx":         6,
        "EBITDA":       7,
    }

    labels = {
        "Revenue": "Revenue",
        "COGS_Pct": "COGS %",
        "COGS": "COGS",
        "Gross_Profit": "Gross Profit",
        "OpEx": "Operating Expenses",
        "EBITDA": "EBITDA",
    }

    for key, row in calc_rows.items():
        ws_c.cell(row=row, column=1, value=labels[key])

    # Named ranges -> cell locations in Calculations sheet
    calc_names = {
        # Revenue
        "Revenue_Y1":       ("B2", "=Inp_Rev_Y1",                          "Year 1 revenue sourced from assumption input"),
        "Revenue_Y2":       ("C2", "=Revenue_Y1*(1+Inp_Rev_Growth)",       "Year 2 revenue = prior year * (1 + growth rate)"),
        "Revenue_Y3":       ("D2", "=Revenue_Y2*(1+Inp_Rev_Growth)",       "Year 3 revenue = prior year * (1 + growth rate)"),
        # COGS %
        "COGS_Pct_Y1":     ("B3", "=Inp_COGS_Pct_Y1",                     "Year 1 COGS percentage from assumption"),
        "COGS_Pct_Y2":     ("C3", "=COGS_Pct_Y1-Inp_COGS_Efficiency",     "Year 2 COGS % reduced by efficiency gain"),
        "COGS_Pct_Y3":     ("D3", "=COGS_Pct_Y2-Inp_COGS_Efficiency",     "Year 3 COGS % reduced by efficiency gain"),
        # COGS $
        "COGS_Y1":         ("B4", "=Revenue_Y1*COGS_Pct_Y1",              "Year 1 COGS = Revenue * COGS percentage"),
        "COGS_Y2":         ("C4", "=Revenue_Y2*COGS_Pct_Y2",              "Year 2 COGS = Revenue * COGS percentage"),
        "COGS_Y3":         ("D4", "=Revenue_Y3*COGS_Pct_Y3",              "Year 3 COGS = Revenue * COGS percentage"),
        # Gross Profit
        "Gross_Profit_Y1": ("B5", "=Revenue_Y1-COGS_Y1",                  "Year 1 Gross Profit = Revenue minus COGS"),
        "Gross_Profit_Y2": ("C5", "=Revenue_Y2-COGS_Y2",                  "Year 2 Gross Profit = Revenue minus COGS"),
        "Gross_Profit_Y3": ("D5", "=Revenue_Y3-COGS_Y3",                  "Year 3 Gross Profit = Revenue minus COGS"),
        # OpEx
        "OpEx_Y1":         ("B6", "=Inp_OpEx_Y1",                          "Year 1 OpEx sourced from assumption input"),
        "OpEx_Y2":         ("C6", "=OpEx_Y1*(1+Inp_OpEx_Growth)",          "Year 2 OpEx = prior year * (1 + growth rate)"),
        "OpEx_Y3":         ("D6", "=OpEx_Y2*(1+Inp_OpEx_Growth)",          "Year 3 OpEx = prior year * (1 + growth rate)"),
        # EBITDA
        "EBITDA_Y1":       ("B7", "=Gross_Profit_Y1-OpEx_Y1",             "Year 1 EBITDA = Gross Profit minus OpEx"),
        "EBITDA_Y2":       ("C7", "=Gross_Profit_Y2-OpEx_Y2",             "Year 2 EBITDA = Gross Profit minus OpEx"),
        "EBITDA_Y3":       ("D7", "=Gross_Profit_Y3-OpEx_Y3",             "Year 3 EBITDA = Gross Profit minus OpEx"),
    }

    for name, (cell, formula, comment_text) in calc_names.items():
        ws_c[cell] = formula
        ws_c[cell].comment = Comment(comment_text, "IDFA Model Builder")
        ws_c[cell].number_format = '#,##0'
        add_named_range(wb, name, "Calculations", cell)

    # ── Sheet: Output ────────────────────────────────────────────────────
    ws_o = wb.create_sheet("Output")
    ws_o.column_dimensions["A"].width = 22
    ws_o.column_dimensions["B"].width = 18
    ws_o.column_dimensions["C"].width = 18
    ws_o.column_dimensions["D"].width = 18

    style_header(ws_o, 1, [1, 2, 3, 4], ["Metric", "Year 1", "Year 2", "Year 3"])

    output_rows = [
        ("Revenue",      "=Revenue_Y1",      "=Revenue_Y2",      "=Revenue_Y3"),
        ("COGS",         "=COGS_Y1",         "=COGS_Y2",         "=COGS_Y3"),
        ("Gross Profit", "=Gross_Profit_Y1", "=Gross_Profit_Y2", "=Gross_Profit_Y3"),
        ("OpEx",         "=OpEx_Y1",         "=OpEx_Y2",         "=OpEx_Y3"),
        ("EBITDA",       "=EBITDA_Y1",       "=EBITDA_Y2",       "=EBITDA_Y3"),
    ]

    for row_idx, (label, y1, y2, y3) in enumerate(output_rows, start=2):
        ws_o.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_o.cell(row=row_idx, column=2, value=y1).number_format = '#,##0'
        ws_o.cell(row=row_idx, column=3, value=y2).number_format = '#,##0'
        ws_o.cell(row=row_idx, column=4, value=y3).number_format = '#,##0'

    path = f"{BASE_DIR_FA}/idfa_compliant_model.xlsx"
    wb.save(path)
    print(f"[OK] Saved: {path}")
    return wb


# ═════════════════════════════════════════════════════════════════════════════
# FILE 2: Legacy (Non-Compliant) Model
# ═════════════════════════════════════════════════════════════════════════════

def create_legacy_model():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # ── Input assumptions (no named ranges — violation) ──────────────────
    ws["A1"] = "Assumptions"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Revenue"
    ws["B2"] = 10_000_000
    ws["A3"] = "Revenue Growth"
    ws["B3"] = 0.10
    ws["A4"] = "COGS %"
    ws["B4"] = 0.60

    # ── Calculations using cell references (violation: no named ranges) ──
    ws["A5"] = ""
    ws["A6"] = "Revenue"
    ws["B6"] = "Y1"
    ws["C6"] = "=B2"                     # cell reference, no named range
    ws["D6"] = "=C6*(1+B3)"             # cell references
    ws["E6"] = "=D6*(1+B3)"             # cell references

    ws["A7"] = "COGS"
    ws["B7"] = "Y1"
    ws["C7"] = "=C6*B4"                 # cell reference
    ws["D7"] = "=D6*0.59"               # HARDCODED VALUE - major violation
    ws["E7"] = "=E6*0.58"               # HARDCODED VALUE - major violation

    ws["A8"] = "Gross Profit"
    ws["B8"] = "Y1"
    ws["C8"] = "=C6-C7"
    ws["D8"] = "=D6-D7"
    ws["E8"] = "=E6-E7"

    # ── WACC section (complex cell-reference formula — violation) ────────
    ws["A10"] = "Equity"
    ws["B10"] = 5_000_000
    ws["A11"] = "Debt"
    ws["B11"] = 3_000_000
    ws["A12"] = "Total Capital"
    ws["B12"] = 8_000_000
    ws["A13"] = "Cost of Equity"
    ws["B13"] = 0.12
    ws["A14"] = "Cost of Debt"
    ws["B14"] = 0.06
    ws["A15"] = "Tax Rate"
    ws["B15"] = 0.25

    ws["A16"] = "WACC"
    ws["B16"] = "=($B$10/$B$12)*$B$13+($B$11/$B$12)*$B$14*(1-$B$15)"

    # Format numbers
    for cell_ref in ["B2", "B10", "B11", "B12"]:
        ws[cell_ref].number_format = '#,##0'
    for cell_ref in ["B3", "B4", "B13", "B14", "B15"]:
        ws[cell_ref].number_format = '0.00%'
    ws["B16"].number_format = '0.00%'
    for col in ["C", "D", "E"]:
        for row in [6, 7, 8]:
            ws[f"{col}{row}"].number_format = '#,##0'

    # NO comments, NO named ranges — deliberate IDFA violations

    path = f"{BASE_DIR_FA}/legacy_model.xlsx"
    wb.save(path)
    print(f"[OK] Saved: {path}")
    return wb


# ═════════════════════════════════════════════════════════════════════════════
# FILE 3: Ops Test Model
# ═════════════════════════════════════════════════════════════════════════════

def create_ops_test_model():
    wb = openpyxl.Workbook()

    # ── Sheet: Assumptions ───────────────────────────────────────────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.column_dimensions["A"].width = 35
    ws_a.column_dimensions["B"].width = 18

    style_header(ws_a, 1, [1, 2], ["Assumption", "Value"])

    assumptions = [
        ("Inp_Price",                  "Unit Price",                100,         "B2"),
        ("Inp_Units_Y1",              "Units Sold Year 1",         50_000,      "B3"),
        ("Inp_Units_Growth",          "Units Growth Rate",          0.15,        "B4"),
        ("Inp_Variable_Cost_Per_Unit", "Variable Cost Per Unit",    40,          "B5"),
        ("Inp_Fixed_Costs",           "Fixed Costs",                1_000_000,   "B6"),
        ("Inp_Tax_Rate",              "Tax Rate",                   0.30,        "B7"),
    ]

    for idx, (name, label, value, cell) in enumerate(assumptions, start=2):
        ws_a.cell(row=idx, column=1, value=label)
        ws_a.cell(row=idx, column=2, value=value)
        if isinstance(value, float) and value < 1:
            ws_a[cell].number_format = '0.00%'
        else:
            ws_a[cell].number_format = '#,##0'
        add_named_range(wb, name, "Assumptions", cell)

    # ── Sheet: Calculations ──────────────────────────────────────────────
    ws_c = wb.create_sheet("Calculations")
    ws_c.column_dimensions["A"].width = 28
    ws_c.column_dimensions["B"].width = 20
    ws_c.column_dimensions["C"].width = 20

    style_header(ws_c, 1, [1, 2, 3], ["Metric", "Year 1", "Year 2"])

    # Layout:
    # Row 2: Units
    # Row 3: Revenue
    # Row 4: Variable Costs
    # Row 5: Contribution
    # Row 6: EBIT
    # Row 7: Tax
    # Row 8: Net Income

    row_labels = {
        2: "Units Sold",
        3: "Revenue",
        4: "Variable Costs",
        5: "Contribution Margin",
        6: "EBIT",
        7: "Tax",
        8: "Net Income",
    }

    for row, label in row_labels.items():
        ws_c.cell(row=row, column=1, value=label)

    calc_names = {
        # Units
        "Units_Y1":           ("B2", "=Inp_Units_Y1",                                "Year 1 units from assumption input"),
        "Units_Y2":           ("C2", "=Units_Y1*(1+Inp_Units_Growth)",                "Year 2 units = prior year * (1 + growth)"),
        # Revenue
        "Revenue_Y1":         ("B3", "=Units_Y1*Inp_Price",                           "Year 1 revenue = units * price per unit"),
        "Revenue_Y2":         ("C3", "=Units_Y2*Inp_Price",                           "Year 2 revenue = units * price per unit"),
        # Variable Costs
        "Variable_Costs_Y1":  ("B4", "=Units_Y1*Inp_Variable_Cost_Per_Unit",          "Year 1 variable costs = units * cost per unit"),
        "Variable_Costs_Y2":  ("C4", "=Units_Y2*Inp_Variable_Cost_Per_Unit",          "Year 2 variable costs = units * cost per unit"),
        # Contribution
        "Contribution_Y1":    ("B5", "=Revenue_Y1-Variable_Costs_Y1",                 "Year 1 contribution = revenue minus variable costs"),
        "Contribution_Y2":    ("C5", "=Revenue_Y2-Variable_Costs_Y2",                 "Year 2 contribution = revenue minus variable costs"),
        # EBIT
        "EBIT_Y1":            ("B6", "=Contribution_Y1-Inp_Fixed_Costs",              "Year 1 EBIT = contribution minus fixed costs"),
        "EBIT_Y2":            ("C6", "=Contribution_Y2-Inp_Fixed_Costs",              "Year 2 EBIT = contribution minus fixed costs"),
        # Tax
        "Tax_Y1":             ("B7", "=EBIT_Y1*Inp_Tax_Rate",                         "Year 1 tax = EBIT * tax rate"),
        "Tax_Y2":             ("C7", "=EBIT_Y2*Inp_Tax_Rate",                         "Year 2 tax = EBIT * tax rate"),
        # Net Income
        "Net_Income_Y1":      ("B8", "=EBIT_Y1-Tax_Y1",                               "Year 1 net income = EBIT minus tax"),
        "Net_Income_Y2":      ("C8", "=EBIT_Y2-Tax_Y2",                               "Year 2 net income = EBIT minus tax"),
    }

    for name, (cell, formula, comment_text) in calc_names.items():
        ws_c[cell] = formula
        ws_c[cell].comment = Comment(comment_text, "IDFA Model Builder")
        ws_c[cell].number_format = '#,##0'
        add_named_range(wb, name, "Calculations", cell)

    # ── Sheet: Output ────────────────────────────────────────────────────
    ws_o = wb.create_sheet("Output")
    ws_o.column_dimensions["A"].width = 28
    ws_o.column_dimensions["B"].width = 20
    ws_o.column_dimensions["C"].width = 20

    style_header(ws_o, 1, [1, 2, 3], ["Metric", "Year 1", "Year 2"])

    output_rows = [
        ("Units Sold",          "=Units_Y1",          "=Units_Y2"),
        ("Revenue",             "=Revenue_Y1",        "=Revenue_Y2"),
        ("Variable Costs",      "=Variable_Costs_Y1", "=Variable_Costs_Y2"),
        ("Contribution Margin", "=Contribution_Y1",   "=Contribution_Y2"),
        ("EBIT",                "=EBIT_Y1",           "=EBIT_Y2"),
        ("Tax",                 "=Tax_Y1",            "=Tax_Y2"),
        ("Net Income",          "=Net_Income_Y1",     "=Net_Income_Y2"),
    ]

    for row_idx, (label, y1, y2) in enumerate(output_rows, start=2):
        ws_o.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_o.cell(row=row_idx, column=2, value=y1).number_format = '#,##0'
        ws_o.cell(row=row_idx, column=3, value=y2).number_format = '#,##0'

    path = f"{BASE_DIR_OPS}/ops_test_model.xlsx"
    wb.save(path)
    print(f"[OK] Saved: {path}")
    return wb


# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("Generating IDFA eval XLSX files...\n")

    wb1 = create_compliant_model()
    wb2 = create_legacy_model()
    wb3 = create_ops_test_model()

    # ── Verification ─────────────────────────────────────────────────────
    print("\n── Verification ──")

    # Check File 1
    print(f"\n[File 1] idfa_compliant_model.xlsx")
    print(f"  Sheets: {wb1.sheetnames}")
    all_names_1 = list(wb1.defined_names.values())
    print(f"  Named ranges ({len(all_names_1)}):")
    for dn in sorted(all_names_1, key=lambda d: d.name):
        print(f"    {dn.name:25s} -> {dn.attr_text}")
    ws_c1 = wb1["Calculations"]
    print(f"  Sample formulas:")
    print(f"    Revenue_Y1 (B2): {ws_c1['B2'].value}")
    print(f"    Revenue_Y2 (C2): {ws_c1['C2'].value}")
    print(f"    EBITDA_Y3  (D7): {ws_c1['D7'].value}")
    print(f"  Comments present: {ws_c1['B2'].comment is not None}")

    # Check File 2
    print(f"\n[File 2] legacy_model.xlsx")
    print(f"  Sheets: {wb2.sheetnames}")
    print(f"  Named ranges: {len(list(wb2.defined_names.values()))}")
    ws_m = wb2["Model"]
    print(f"  WACC formula (B16): {ws_m['B16'].value}")
    print(f"  D7 (hardcoded):     {ws_m['D7'].value}")
    print(f"  E7 (hardcoded):     {ws_m['E7'].value}")
    print(f"  Comments present on C6: {ws_m['C6'].comment is not None}")

    # Check File 3
    print(f"\n[File 3] ops_test_model.xlsx")
    print(f"  Sheets: {wb3.sheetnames}")
    all_names_3 = list(wb3.defined_names.values())
    print(f"  Named ranges ({len(all_names_3)}):")
    for dn in sorted(all_names_3, key=lambda d: d.name):
        print(f"    {dn.name:30s} -> {dn.attr_text}")
    ws_c3 = wb3["Calculations"]
    print(f"  Sample formulas:")
    print(f"    Revenue_Y1 (B3): {ws_c3['B3'].value}")
    print(f"    Net_Income_Y2 (C8): {ws_c3['C8'].value}")
    print(f"  Comments present: {ws_c3['B3'].comment is not None}")

    print("\nDone.")
